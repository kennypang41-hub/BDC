"""Build the hyperscaler commitments workbook from ``data/commitments.json``.

One row per (company, quarter, metric) on the ``Data`` sheet. The matrix sheets
compute from it with SUMIFS rather than baked-in numbers, so correcting a row on
``Data`` updates every view. Cells with no disclosure stay genuinely empty --
a missing figure is never rendered as a zero.
"""
from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent
DATA = ROOT / "data" / "commitments.json"
OUT = ROOT / "output" / "hyperscaler_commitments.xlsx"

FONT = "Arial"
BN = '$#,##0.0,,"bn";($#,##0.0,,"bn");-'   # values stored in $m, shown as $bn
MONEY_M = "#,##0"

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(name=FONT, bold=True, size=13)
BODY = Font(name=FONT, size=10)
MUTED = Font(name=FONT, size=9, color="666666")
TOTAL_FONT = Font(name=FONT, bold=True, size=10)
TOTAL_FILL = PatternFill("solid", fgColor="EDF1F8")
THIN = Side(style="thin", color="BFBFBF")

METRICS = [
    ("uncommenced_leases", "Uncommenced lease commitments"),
    ("purchase_commitments", "Purchase commitments"),
]


def load() -> dict:
    return json.loads(DATA.read_text())


def style_header(ws, row: int, ncols: int, start: int = 1) -> None:
    for col in range(start, start + ncols):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=THIN)
    ws.row_dimensions[row].height = 28


def build_readme(wb: Workbook, d: dict) -> None:
    ws = wb.create_sheet("Read me")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 108

    rows = [
        ("Hyperscaler off-balance-sheet commitments", None, TITLE_FONT),
        (None, None, None),
        ("What this is",
         "Twelve quarters of two footnote disclosures for six hyperscalers: signed leases that have not yet "
         "commenced, and non-cancelable purchase commitments. Both sit outside the balance sheet but bind future cash.", None),
        ("Units", "All figures in USD millions on the Data sheet; matrix sheets display $bn.", None),
        ("Quarters", f"{d['quarters'][0]} to {d['quarters'][-1]} (calendar-aligned).", None),
        (None, None, None),
        ("Why leases sit off balance sheet",
         d["meta"]["metrics"]["uncommenced_leases"], None),
        ("Purchase commitments",
         d["meta"]["metrics"]["purchase_commitments"], None),
        (None, None, None),
        ("Definitions are NOT uniform", d["series_notes"]["definitions"], None),
        (None, None, None),
        ("How the figures were collected", d["meta"]["collection_note"], None),
        (None, None, None),
        ("Provenance codes", None, TOTAL_FONT),
    ]
    for label, text, font in rows:
        r = ws.max_row + 1 if ws.max_row > 1 or label else 1
        ws.cell(row=r, column=1, value=label).font = font or Font(name=FONT, bold=True, size=10)
        if text:
            c = ws.cell(row=r, column=2, value=text)
            c.font = BODY
            c.alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[r].height = max(15, 13 * (len(text) // 105 + 1))

    for code, desc in d["meta"]["provenance"].items():
        r = ws.max_row + 1
        ws.cell(row=r, column=1, value=code).font = BODY
        c = ws.cell(row=r, column=2, value=desc)
        c.font = BODY
        c.alignment = Alignment(wrap_text=True, vertical="top")

    r = ws.max_row + 2
    ws.cell(row=r, column=1, value="Series with no data").font = TOTAL_FONT
    for key in ("NVDA_uncommenced_leases", "ORCL_early"):
        r = ws.max_row + 1
        ws.cell(row=r, column=1, value=key).font = BODY
        c = ws.cell(row=r, column=2, value=d["series_notes"][key])
        c.font = BODY
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = max(15, 13 * (len(d["series_notes"][key]) // 105 + 1))

    r = ws.max_row + 2
    c = ws.cell(row=r, column=1, value="Empty cell = the company did not disclose a figure we could source for that quarter. "
                                       "It is not a zero, and it is not an estimate.")
    c.font = Font(name=FONT, bold=True, size=10, color="9C2B2B")


def build_data(wb: Workbook, d: dict) -> str:
    ws = wb.create_sheet("Data")
    ws.freeze_panes = "A2"
    headers = ["Ticker", "Company", "CIK", "Calendar quarter", "Period end",
               "Metric", "Metric label", "Value ($m)", "Value ($bn)", "Provenance", "Note"]
    widths = [9, 24, 12, 16, 12, 22, 30, 14, 12, 13, 90]
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        ws.cell(row=1, column=i, value=h)
        ws.column_dimensions[get_column_letter(i)].width = w
    style_header(ws, 1, len(headers))

    names = {c["ticker"]: c for c in d["companies"]}
    labels = dict(METRICS)
    order = {q: i for i, q in enumerate(d["quarters"])}
    obs = sorted(d["observations"],
                 key=lambda o: (o["metric"], names[o["ticker"]]["slot"], order[o["quarter"]]))

    for row, o in enumerate(obs, start=2):
        comp = names[o["ticker"]]
        ws.cell(row=row, column=1, value=o["ticker"])
        ws.cell(row=row, column=2, value=comp["name"])
        ws.cell(row=row, column=3, value=comp["cik"])
        ws.cell(row=row, column=4, value=o["quarter"])
        ws.cell(row=row, column=5, value=d["period_ends"][o["ticker"]][o["quarter"]])
        ws.cell(row=row, column=6, value=o["metric"])
        ws.cell(row=row, column=7, value=labels[o["metric"]])
        ws.cell(row=row, column=8, value=round(o["value"] * 1000, 1)).number_format = MONEY_M
        ws.cell(row=row, column=9, value=f"=H{row}/1000").number_format = "#,##0.0"
        ws.cell(row=row, column=10, value=o["provenance"])
        ws.cell(row=row, column=11, value=o["note"])
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).font = BODY
        ws.cell(row=row, column=11).alignment = Alignment(wrap_text=False)

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(obs) + 1}"
    return f"Data!$A$2:$A${len(obs) + 1}", len(obs) + 1


def build_matrix(wb: Workbook, d: dict, metric: str, title: str, last_row: int) -> None:
    ws = wb.create_sheet(title[:31])
    ws.sheet_view.showGridLines = False
    tickers = [c["ticker"] for c in d["companies"]]

    ws.cell(row=1, column=1, value=title).font = TITLE_FONT
    ws.cell(row=2, column=1, value="USD billions. Blank = not disclosed / not sourced; never treated as zero.").font = MUTED

    ws.column_dimensions["A"].width = 18
    ws.cell(row=4, column=1, value="Calendar quarter")
    for i, t in enumerate(tickers, start=2):
        ws.cell(row=4, column=i, value=t)
        ws.column_dimensions[get_column_letter(i)].width = 13
    ws.cell(row=4, column=len(tickers) + 2, value="Total")
    ws.column_dimensions[get_column_letter(len(tickers) + 2)].width = 14
    ws.cell(row=4, column=len(tickers) + 3, value="Reported")
    ws.column_dimensions[get_column_letter(len(tickers) + 3)].width = 11
    style_header(ws, 4, len(tickers) + 3)

    dr = f"$H$2:$H${last_row}"
    tr = f"$A$2:$A${last_row}"
    qr = f"$D$2:$D${last_row}"
    mr = f"$F$2:$F${last_row}"

    first = 5
    for r, q in enumerate(d["quarters"], start=first):
        ws.cell(row=r, column=1, value=q).font = BODY
        for i, t in enumerate(tickers, start=2):
            crit = (f'Data!{tr},"{t}",Data!{qr},$A{r},Data!{mr},"{metric}"')
            f = f'=IF(COUNTIFS({crit})=0,"",SUMIFS(Data!{dr},{crit})/1000)'
            c = ws.cell(row=r, column=i, value=f)
            c.number_format = "#,##0.0"
            c.font = BODY
            c.border = Border(bottom=THIN)
        span = f"{get_column_letter(2)}{r}:{get_column_letter(len(tickers) + 1)}{r}"
        tot = ws.cell(row=r, column=len(tickers) + 2, value=f'=IF(COUNT({span})=0,"",SUM({span}))')
        tot.number_format = "#,##0.0"
        tot.font = TOTAL_FONT
        tot.fill = TOTAL_FILL
        cnt = ws.cell(row=r, column=len(tickers) + 3, value=f'=COUNT({span})&" of 6"')
        cnt.font = MUTED
        cnt.alignment = Alignment(horizontal="center")

    last = first + len(d["quarters"]) - 1
    r = last + 2
    ws.cell(row=r, column=1, value="Change over the period").font = TOTAL_FONT
    for i, t in enumerate(tickers, start=2):
        col = get_column_letter(i)
        f = (f'=IF(OR({col}{last}="",COUNT({col}{first}:{col}{last})<2),"",'
             f'{col}{last}-INDEX({col}{first}:{col}{last},MATCH(TRUE,INDEX({col}{first}:{col}{last}<>"",0),0)))')
        c = ws.cell(row=r, column=i, value=f)
        c.number_format = '+#,##0.0;-#,##0.0;-'
        c.font = TOTAL_FONT
    ws.cell(row=r + 1, column=1,
            value="Latest less the first quarter that company disclosed, so it is not comparable across columns "
                  "with different start dates.").font = MUTED


def build_combined(wb: Workbook, d: dict) -> None:
    ws = wb.create_sheet("Combined")
    ws.sheet_view.showGridLines = False
    tickers = [c["ticker"] for c in d["companies"]]
    leases, purchases = (m[1] for m in METRICS)

    ws.cell(row=1, column=1, value="Uncommenced leases + purchase commitments").font = TITLE_FONT
    ws.cell(row=2, column=1,
            value="USD billions. A company is summed only where both legs exist, so a blank here means at least "
                  "one leg is missing for that quarter.").font = MUTED
    ws.column_dimensions["A"].width = 18
    ws.cell(row=4, column=1, value="Calendar quarter")
    for i, t in enumerate(tickers, start=2):
        ws.cell(row=4, column=i, value=t)
        ws.column_dimensions[get_column_letter(i)].width = 13
    ws.cell(row=4, column=len(tickers) + 2, value="Total")
    ws.column_dimensions[get_column_letter(len(tickers) + 2)].width = 14
    style_header(ws, 4, len(tickers) + 1)
    style_header(ws, 4, 1, start=len(tickers) + 2)

    for r, q in enumerate(d["quarters"], start=5):
        ws.cell(row=r, column=1, value=q).font = BODY
        for i in range(2, len(tickers) + 2):
            col = get_column_letter(i)
            f = (f"=IF(AND('{leases}'!{col}{r}<>\"\",'{purchases}'!{col}{r}<>\"\"),"
                 f"'{leases}'!{col}{r}+'{purchases}'!{col}{r},\"\")")
            c = ws.cell(row=r, column=i, value=f)
            c.number_format = "#,##0.0"
            c.font = BODY
            c.border = Border(bottom=THIN)
        span = f"B{r}:{get_column_letter(len(tickers) + 1)}{r}"
        tot = ws.cell(row=r, column=len(tickers) + 2, value=f'=IF(COUNT({span})=0,"",SUM({span}))')
        tot.number_format = "#,##0.0"
        tot.font = TOTAL_FONT
        tot.fill = TOTAL_FILL


def build_coverage(wb: Workbook, d: dict) -> None:
    ws = wb.create_sheet("Coverage")
    ws.sheet_view.showGridLines = False
    ws.cell(row=1, column=1, value="What is populated, and how well sourced").font = TITLE_FONT
    ws.cell(row=2, column=1,
            value="Read this before using the numbers. It is the honest map of where the dataset is solid "
                  "and where it is thin.").font = MUTED

    have = {(o["ticker"], o["quarter"], o["metric"]): o for o in d["observations"]}
    headers = ["Ticker", "Metric", "Quarters populated", "of 12", "filing / filing_table", "press / derived"]
    for i, (h, w) in enumerate(zip(headers, [9, 30, 19, 8, 22, 18]), start=1):
        ws.cell(row=4, column=i, value=h)
        ws.column_dimensions[get_column_letter(i)].width = w
    style_header(ws, 4, len(headers))

    r = 5
    for comp in d["companies"]:
        for metric, label in METRICS:
            rows = [have[k] for k in have if k[0] == comp["ticker"] and k[2] == metric]
            strong = sum(1 for o in rows if o["provenance"] in ("filing", "filing_table"))
            weak = len(rows) - strong
            ws.cell(row=r, column=1, value=comp["ticker"]).font = BODY
            ws.cell(row=r, column=2, value=label).font = BODY
            ws.cell(row=r, column=3, value=len(rows)).font = BODY
            ws.cell(row=r, column=4, value=12).font = MUTED
            ws.cell(row=r, column=5, value=strong).font = BODY
            ws.cell(row=r, column=6, value=weak).font = BODY
            r += 1

    r += 1
    ws.cell(row=r, column=1,
            value="2023 and most of 2024 are thin by nature: several of these companies only began disclosing a "
                  "material uncommenced-lease balance during 2025.").font = MUTED


def main() -> None:
    d = load()
    wb = Workbook()
    wb.remove(wb.active)
    build_readme(wb, d)
    _, last_row = build_data(wb, d)
    for metric, label in METRICS:
        build_matrix(wb, d, metric, label, last_row)
    build_combined(wb, d)
    build_coverage(wb, d)
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"wrote {OUT} ({len(d['observations'])} observations)")


if __name__ == "__main__":
    main()
