"""Export the mark dataset to a workbook.

One row per (loan, quarter) on the ``Marks`` sheet — the same grain as the
database — plus summary sheets that compute from it with live formulas rather
than baked-in numbers, so filtering or correcting a row updates the totals.
"""
from __future__ import annotations

import sqlite3
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from bdctracker import analytics

FONT = "Arial"
MONEY = "$#,##0;($#,##0);-"
MONEY_SIGNED = "$#,##0;[Red]($#,##0);-"
PCT = "0.00%"
PCT1 = "0.0%"
PRICE = "0.0"
INT = "#,##0"

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=10)
WARN_FILL = PatternFill("solid", fgColor="FFFF00")
THIN = Side(style="thin", color="BFBFBF")

#: (header, source column or None, number format, width). ``None`` marks a
#: column filled with a per-row formula rather than a stored value.
MARK_COLUMNS: list[tuple[str, str | None, str | None, int]] = [
    ("Period end",            "period_end",        None,          12),
    ("BDC",                   "ticker",            None,          8),
    ("BDC name",              "bdc_name",          None,          30),
    ("CIK",                   "cik",               "0",           10),
    ("Borrower",              "issuer_name",       None,          38),
    ("Instrument",            "investment_type",   None,          18),
    ("Lien",                  "lien",              None,          12),
    ("Facility",              "facility",          None,          12),
    ("Industry",              "industry",          None,          26),
    ("Ccy",                   "currency",          None,          6),
    ("Principal",             "principal",         MONEY,         16),
    ("Cost",                  "cost",              MONEY,         16),
    ("Fair value",            "fair_value",        MONEY,         16),
    ("Mark basis (fallback)", "mark_basis",        MONEY,         16),
    ("Mark (% of principal)", None,                PRICE,         14),
    ("Unrealised",            None,                MONEY_SIGNED,  16),
    ("Unrealised vs cost",    None,                PCT1,          16),
    ("Coupon",                "interest_rate",     PCT,           10),
    ("Spread",                "spread",            PCT,           10),
    ("Base rate",             "reference_rate",    None,          12),
    ("PIK",                   "pik_rate",          PCT,           10),
    ("% of net assets",       "pct_net_assets",    PCT,           14),
    ("Maturity",              "maturity_date",     None,          12),
    ("Acquired",              "acquisition_date",  None,          12),
    ("FV level",              "fair_value_level",  None,          10),
    ("Non-accrual",           "non_accrual_text",  None,          12),
    ("Debt",                  "is_debt",           "0",           6),
    ("Shares",                "shares",            INT,           14),
    ("Loan id",               "loan_id",           None,          22),
    ("Borrower id",           "issuer_id",         None,          22),
    ("Credit id",             "credit_id",         None,          22),
    ("Accession",             "accession",         None,          22),
    ("Form",                  "form",              None,          8),
    ("Filed",                 "filed_date",        None,          12),
    ("Source",                "source",            None,          10),
    ("Flags",                 "flags",             None,          26),
]

_COLUMN_INDEX = {header: i + 1 for i, (header, *_rest) in enumerate(MARK_COLUMNS)}


def _letter(header: str) -> str:
    return get_column_letter(_COLUMN_INDEX[header])


MARKS_SQL = """
SELECT
    m.period_end, b.ticker, b.name AS bdc_name, m.cik,
    i.display_name AS issuer_name,
    l.investment_type, l.lien, l.facility, m.industry, l.currency,
    m.principal, m.cost, m.fair_value, m.shares,
    CASE WHEN m.principal > 0 THEN m.principal ELSE m.cost END AS mark_basis,
    m.interest_rate, m.spread, m.reference_rate, m.pik_rate, m.pct_net_assets,
    m.maturity_date, m.acquisition_date, m.fair_value_level,
    CASE WHEN m.is_non_accrual = 1 THEN 'Yes'
         WHEN m.is_non_accrual = 0 THEN 'No'
         ELSE 'Not disclosed' END AS non_accrual_text,
    l.is_debt, m.loan_id, m.issuer_id, m.credit_id,
    m.accession, m.form, m.filed_date, m.source, m.flags
FROM marks m
JOIN loans   l ON l.loan_id  = m.loan_id
JOIN bdcs    b ON b.cik      = m.cik
JOIN issuers i ON i.issuer_id = m.issuer_id
ORDER BY m.period_end DESC, b.ticker, m.fair_value DESC
"""


def _rate(value):
    """DB stores rates as percentage points; Excel percentages are fractions."""
    return None if value is None else value / 100.0


def _style_header(sheet: Worksheet, columns: list[tuple]) -> None:
    for index, (header, _src, _fmt, width) in enumerate(columns, start=1):
        cell = sheet.cell(row=1, column=index, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = "A2"


def _write_marks(workbook: Workbook, conn: sqlite3.Connection) -> tuple[int, dict[str, tuple[int, int]]]:
    """Write the fact table; return the row count and each period's row block.

    The query orders by period, so every period occupies a contiguous run of
    rows. Handing those bounds to the summary sheet lets its SUMIFS scan one
    quarter instead of the whole history — the difference between a workbook
    that opens instantly and one that recalculates for minutes.
    """
    sheet = workbook.create_sheet("Marks")
    _style_header(sheet, MARK_COLUMNS)

    fair_value = _letter("Fair value")
    basis = _letter("Mark basis (fallback)")
    principal = _letter("Principal")
    cost = _letter("Cost")
    mark = _letter("Mark (% of principal)")
    unrealised = _letter("Unrealised")

    blocks: dict[str, list[int]] = {}
    row_number = 1
    for record in conn.execute(MARKS_SQL):
        row_number += 1
        row = dict(record)
        period = row["period_end"]
        if period in blocks:
            blocks[period][1] = row_number
        else:
            blocks[period] = [row_number, row_number]
        values = []
        for header, source, number_format, _width in MARK_COLUMNS:
            if source is None:
                if header == "Mark (% of principal)":
                    # Principal is the denominator; mark basis only stands in
                    # where the filing reports no principal (equity, or debt the
                    # filer left untagged).
                    value = (
                        f"=IFERROR({fair_value}{row_number}"
                        f"/IF({principal}{row_number}>0,{principal}{row_number},"
                        f"{basis}{row_number})*100,\"\")"
                    )
                elif header == "Unrealised":
                    value = f"=IFERROR({fair_value}{row_number}-{cost}{row_number},\"\")"
                else:  # Unrealised vs cost
                    value = f"=IFERROR({unrealised}{row_number}/{cost}{row_number},\"\")"
            elif number_format == PCT:
                value = _rate(row.get(source))
            else:
                value = row.get(source)
            values.append(value)

        sheet.append(values)
        for index, (_header, _source, number_format, _width) in enumerate(MARK_COLUMNS, start=1):
            if number_format:
                sheet.cell(row=row_number, column=index).number_format = number_format

    sheet.auto_filter.ref = f"A1:{get_column_letter(len(MARK_COLUMNS))}{row_number}"
    # Highlight the mark itself — it is what the whole workbook is for.
    for cell in sheet[mark][1:]:
        cell.font = Font(name=FONT, bold=True)
    return row_number - 1, {k: (v[0], v[1]) for k, v in blocks.items()}


SUMMARY_COLUMNS = [
    ("BDC", None, None, 8),
    ("Name", None, None, 32),
    ("Positions", None, INT, 12),
    ("Principal", None, MONEY, 16),
    ("Cost", None, MONEY, 16),
    ("Fair value", None, MONEY, 16),
    ("Portfolio mark", None, PRICE, 14),
    ("FV / cost", None, PCT1, 12),
    ("Unrealised", None, MONEY_SIGNED, 16),
    ("Non-accrual FV", None, MONEY, 16),
    ("Non-accrual %", None, PCT1, 14),
    ("Status disclosed", None, INT, 16),
    ("Borrowers", None, INT, 12),
]


def _write_summary(workbook: Workbook, conn: sqlite3.Connection,
                   block: tuple[int, int], period: str) -> None:
    """Per-BDC totals, computed from the Marks sheet with live SUMIFS.

    Ranges cover only this period's rows. The period criterion is kept anyway,
    so a stray row inside the block still cannot be counted twice.
    """
    sheet = workbook.create_sheet("BDC summary")
    _style_header(sheet, SUMMARY_COLUMNS)

    first, last = block
    period_range = f"Marks!${_letter('Period end')}${first}:${_letter('Period end')}${last}"
    ticker_range = f"Marks!${_letter('BDC')}${first}:${_letter('BDC')}${last}"

    def sumifs(column_header: str, row: int, extra: str = "") -> str:
        column = _letter(column_header)
        return (
            f"=SUMIFS(Marks!${column}${first}:${column}${last},"
            f"{period_range},$N$1,{ticker_range},$A{row}{extra})"
        )

    sheet["N1"] = period
    sheet["N1"].font = Font(name=FONT, bold=True)
    sheet["M1"] = "Period:"
    sheet["M1"].font = Font(name=FONT, bold=True)

    rows = conn.execute(
        """
        SELECT b.ticker, b.name, COUNT(DISTINCT m.issuer_id) AS issuers
        FROM marks m JOIN bdcs b ON b.cik = m.cik
        WHERE m.period_end = ?
        GROUP BY b.ticker, b.name ORDER BY SUM(m.fair_value) DESC
        """,
        (period,),
    ).fetchall()

    na_range = f"Marks!${_letter('Non-accrual')}${first}:${_letter('Non-accrual')}${last}"
    debt_range = f"Marks!${_letter('Debt')}${first}:${_letter('Debt')}${last}"
    fv = _letter("Fair value")

    for index, record in enumerate(rows, start=2):
        sheet.cell(row=index, column=1, value=record["ticker"])
        sheet.cell(row=index, column=2, value=record["name"])
        sheet.cell(row=index, column=3,
                   value=f"=COUNTIFS({period_range},$N$1,{ticker_range},$A{index})")
        sheet.cell(row=index, column=4, value=sumifs("Principal", index, f",{debt_range},1"))
        sheet.cell(row=index, column=5, value=sumifs("Cost", index))
        sheet.cell(row=index, column=6, value=sumifs("Fair value", index))
        # Portfolio mark is debt fair value over debt par — equity has no par.
        debt_fv = (f"SUMIFS(Marks!${fv}${first}:${fv}${last},{period_range},$N$1,"
                   f"{ticker_range},$A{index},{debt_range},1)")
        sheet.cell(row=index, column=7, value=f"=IFERROR({debt_fv}/$D{index}*100,\"\")")
        sheet.cell(row=index, column=8, value=f"=IFERROR($F{index}/$E{index},\"\")")
        sheet.cell(row=index, column=9, value=f"=$F{index}-$E{index}")
        # Non-accrual status lives in footnotes. Where a filing disclosed none,
        # 0% would assert a clean book we cannot see, so both cells stay blank
        # and "Status disclosed" says how many rows we actually know about.
        disclosed = (
            f'=COUNTIFS({period_range},$N$1,{ticker_range},$A{index},{na_range},"Yes")'
            f'+COUNTIFS({period_range},$N$1,{ticker_range},$A{index},{na_range},"No")'
        )
        sheet.cell(row=index, column=12, value=disclosed)
        sheet.cell(row=index, column=10,
                   value=f'=IF($L{index}=0,"",'
                         + sumifs("Fair value", index, f',{na_range},"Yes"').lstrip("=")
                         + ")")
        sheet.cell(row=index, column=11,
                   value=f'=IF($L{index}=0,"",IFERROR($J{index}/$F{index},""))')
        sheet.cell(row=index, column=13, value=record["issuers"])

        for column, (_h, _s, number_format, _w) in enumerate(SUMMARY_COLUMNS, start=1):
            cell = sheet.cell(row=index, column=column)
            cell.font = Font(name=FONT)
            if number_format:
                cell.number_format = number_format

    total = len(rows) + 2
    sheet.cell(row=total, column=1, value="Total").font = Font(name=FONT, bold=True)
    for column in (3, 4, 5, 6, 9, 10, 12, 13):
        letter = get_column_letter(column)
        cell = sheet.cell(row=total, column=column,
                          value=f"=SUM({letter}2:{letter}{total - 1})")
        cell.font = Font(name=FONT, bold=True)
        cell.number_format = SUMMARY_COLUMNS[column - 1][2] or INT
        cell.border = Border(top=THIN)
    for column, formula in ((7, f'=IFERROR(F{total}/D{total}*100,"")'),
                            (8, f'=IFERROR(F{total}/E{total},"")'),
                            (11, f'=IF(L{total}=0,"",IFERROR(J{total}/F{total},""))')):
        cell = sheet.cell(row=total, column=column, value=formula)
        cell.font = Font(name=FONT, bold=True)
        cell.number_format = SUMMARY_COLUMNS[column - 1][2]
        cell.border = Border(top=THIN)


def _write_disagreements(workbook: Workbook, conn: sqlite3.Connection, period: str) -> None:
    sheet = workbook.create_sheet("Disagreements")
    columns = [
        ("Borrower", None, None, 38), ("BDCs holding", None, INT, 14),
        ("Lowest mark", None, PRICE, 12), ("Highest mark", None, PRICE, 12),
        ("Spread", None, PRICE, 10), ("Weighted mark", None, PRICE, 14),
        ("Fair value", None, MONEY, 16), ("Marks by BDC", None, None, 70),
    ]
    _style_header(sheet, columns)
    for index, row in enumerate(analytics.disagreements(conn, period, limit=1000), start=2):
        sheet.append([
            row["issuer_name"], row["holders"], row["min_mark"], row["max_mark"],
            row["spread"], row["weighted_mark"], row["fair_value"], row["marks_by_bdc"],
        ])
        for column, (_h, _s, number_format, _w) in enumerate(columns, start=1):
            cell = sheet.cell(row=index, column=column)
            cell.font = Font(name=FONT)
            if number_format:
                cell.number_format = number_format


def _write_readme(workbook: Workbook, conn: sqlite3.Connection, period: str,
                  sources: list[str]) -> None:
    sheet = workbook.create_sheet("Read me", 0)
    sheet.column_dimensions["A"].width = 26
    sheet.column_dimensions["B"].width = 96

    synthetic = sources == ["demo"]
    line = 1

    def put(label: str, value: str = "", bold: bool = False, fill: PatternFill | None = None):
        nonlocal line
        left = sheet.cell(row=line, column=1, value=label)
        right = sheet.cell(row=line, column=2, value=value)
        left.font = Font(name=FONT, bold=True, size=11)
        right.font = Font(name=FONT, bold=bold, size=11)
        right.alignment = Alignment(wrap_text=True, vertical="top")
        if fill:
            left.fill = right.fill = fill
        line += 1

    put("BDC Tracker", "Loan-level valuation marks for US Business Development Companies", bold=True)
    line += 1

    if synthetic:
        put("⚠ WARNING",
            "THIS FILE CONTAINS SYNTHETIC DATA. Every row was generated locally by "
            "`bdc demo` for user-interface development. Nothing here was extracted from "
            "an SEC filing. The borrower names, marks and accession numbers are "
            "fabricated. Do not use for analysis, valuation or distribution.",
            bold=True, fill=WARN_FILL)
        line += 1

    stats = conn.execute(
        """
        SELECT COUNT(*) AS marks, COUNT(DISTINCT loan_id) AS loans,
               COUNT(DISTINCT cik) AS bdcs, COUNT(DISTINCT issuer_id) AS issuers,
               COUNT(DISTINCT period_end) AS periods,
               MIN(period_end) AS first_period, MAX(period_end) AS last_period
        FROM marks
        """
    ).fetchone()

    put("Generated", datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC"))
    put("Data source", ", ".join(sources) or "none")
    put("Marks (rows)", f"{stats['marks']:,}")
    put("Loans", f"{stats['loans']:,}")
    put("BDCs", f"{stats['bdcs']:,}")
    put("Borrowers", f"{stats['issuers']:,}")
    put("Quarters", f"{stats['periods']} ({stats['first_period']} to {stats['last_period']})")
    put("Summary period", period)
    line += 1

    put("What a mark is",
        "Fair value divided by principal, as reported in the filing, times 100. "
        "100 means the BDC carries the loan at par; below 100 means it has written the "
        "position down. Equity has no principal, so it is marked against cost instead — "
        "the 'Mark basis (fallback)' column holds that stand-in denominator.")
    put("Principal",
        "Principal outstanding as the filing reports it. Blank where the filer did not "
        "tag it, in which case the mark falls back to cost and the row carries the "
        "no_principal flag.")
    put("Grain", "One row per (loan, quarter) on the Marks sheet.")
    put("Rates", "Coupon, spread and PIK are true percentages (5.75% is stored as 0.0575).")
    put("Non-accrual",
        "'Not disclosed' is not the same as 'No'. Non-accrual status lives in filing "
        "footnotes; where a filing disclosed none, status is left unknown rather than "
        "asserted as accruing.")
    line += 1

    put("Source detail",
        "dera = SEC DERA quarterly BDC Data Sets; xbrl = per-filing XBRL from the "
        "10-K/10-Q Schedule of Investments; demo = synthetic, not from EDGAR.")
    put("Flags legend",
        "no_fair_value / no_principal = the filing omitted it · implausible_mark = outside "
        "1-200, usually a filer units error · unclassified = the instrument label matched "
        "no rule · merged_N_facilities = N same-kind facilities to one borrower summed "
        "within a filing · dropped_restatement = a sibling row repeated this fair value "
        "with no principal or cost (a fair-value-hierarchy breakdown of money already "
        "counted) and was excluded · rescaled_xN = the filing's units were off by N and "
        "corrected.")
    line += 1
    put("Verify a row",
        "The Accession column is the SEC accession number. Look it up at "
        "https://www.sec.gov/cgi-bin/browse-edgar?action=getcompany and read the "
        "Consolidated Schedule of Investments in that filing.")

    sheet.sheet_view.showGridLines = False


def export_workbook(conn: sqlite3.Connection, path: str | Path,
                    period: str | None = None) -> dict:
    """Write the whole dataset to ``path`` and report what went in."""
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    period = period or analytics.latest_period(conn)
    if period is None:
        raise ValueError("the database has no marks; run `bdc harvest` first")

    sources = sorted(r[0] for r in conn.execute("SELECT DISTINCT source FROM marks") if r[0])

    workbook = Workbook()
    workbook.remove(workbook.active)
    marks_rows, blocks = _write_marks(workbook, conn)
    _write_summary(workbook, conn, blocks[period], period)
    _write_disagreements(workbook, conn, period)
    _write_readme(workbook, conn, period, sources)
    workbook.save(path)

    return {
        "path": str(path),
        "marks": marks_rows,
        "period": period,
        "sources": sources,
        "synthetic": sources == ["demo"],
    }
