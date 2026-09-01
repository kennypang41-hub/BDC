"""The workbook and the standalone page are deliverables — check what they carry."""
from __future__ import annotations

import json
import re
from datetime import date
from decimal import Decimal

import pytest
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from bdctracker import analytics, db, excel, normalize, standalone
from bdctracker.models import Position
from bdctracker.universe import BDC

ARCC = BDC(ticker="ARCC", cik=1287750, name="Ares Capital", exchange="Nasdaq")
TSLX = BDC(ticker="TSLX", cik=1508655, name="Sixth Street", exchange="NYSE")
Q = Q2 = date(2025, 12, 31)


def _position(cik, identifier, fv, par, cost=None, **kwargs):
    return normalize.finalize(
        Position(
            cik=cik, period_end=Q, identifier=identifier, source="dera",
            fair_value=Decimal(str(fv)), principal=Decimal(str(par)),
            cost=Decimal(str(cost if cost is not None else par)),
            accession="0001-25-000001", form="10-K", **kwargs,
        )
    )


@pytest.fixture
def conn():
    connection = db.connect(":memory:")
    db.init_schema(connection)
    db.upsert_bdcs(connection, [ARCC, TSLX])
    db.load_positions(connection, [
        _position(ARCC.cik, "Acme Holdings, LLC, First Lien Term Loan",
                  9_000_000, 10_000_000, interest_rate=11.25, spread=5.75),
        _position(TSLX.cik, "ACME Holdings Inc., First Lien Senior Secured Loan",
                  7_000_000, 10_000_000),
        _position(ARCC.cik, "Beta Industries, Second Lien Term Loan",
                  2_000_000, 5_000_000, is_non_accrual=True),
        # Equity: no par, so the mark basis must fall back to cost.
        _position(ARCC.cik, "Beta Industries, Common Equity", 100_000, 0, cost=500_000),
    ])
    connection.commit()
    yield connection
    connection.close()


def test_workbook_has_every_sheet_and_row(conn, tmp_path):
    target = tmp_path / "marks.xlsx"
    result = excel.export_workbook(conn, target)
    assert result["marks"] == 4
    assert result["synthetic"] is False

    book = load_workbook(target)
    assert book.sheetnames == [
        "Read me", "Marks", "BDC summary",
        "Mark by quarter", "Non-accrual mark", "Non-accrual %",
        "Vintage & maturity", "Disagreements",
    ]
    assert book["Marks"].max_row == 5  # header + 4


def test_mark_is_a_live_formula_not_a_baked_number(conn, tmp_path):
    """Editing a fair value must move the mark, so it cannot be hardcoded."""
    target = tmp_path / "marks.xlsx"
    excel.export_workbook(conn, target)
    sheet = load_workbook(target)["Marks"]

    headers = [cell.value for cell in sheet[1]]
    mark_column = headers.index("Mark (% of principal)") + 1
    formula = sheet.cell(row=2, column=mark_column).value
    assert isinstance(formula, str) and formula.startswith("=IFERROR(")
    assert "*100" in formula

    # Principal is the denominator, but only when it is the same figure as the
    # mark basis — which encodes the currency test.
    principal = get_column_letter(headers.index("Principal") + 1)
    basis = get_column_letter(headers.index("Mark basis (fallback)") + 1)
    assert f"AND({principal}2>0,{principal}2={basis}2)" in formula
    assert f",{principal}2,{basis}2)" in formula


def test_mark_basis_is_par_for_debt_and_cost_for_equity(conn, tmp_path):
    target = tmp_path / "marks.xlsx"
    excel.export_workbook(conn, target)
    sheet = load_workbook(target)["Marks"]
    headers = [cell.value for cell in sheet[1]]
    basis = headers.index("Mark basis (fallback)") + 1
    borrower = headers.index("Borrower") + 1
    instrument = headers.index("Instrument") + 1

    rows = {
        (sheet.cell(r, borrower).value, sheet.cell(r, instrument).value):
            sheet.cell(r, basis).value
        for r in range(2, sheet.max_row + 1)
    }
    assert rows[("ACME Holdings Inc.", "FIRST_LIEN")] == 10_000_000
    assert rows[("Beta Industries", "COMMON_EQUITY")] == 500_000


def test_rates_are_stored_as_fractions_so_excel_renders_them_as_percentages(conn, tmp_path):
    target = tmp_path / "marks.xlsx"
    excel.export_workbook(conn, target)
    sheet = load_workbook(target)["Marks"]
    headers = [cell.value for cell in sheet[1]]
    coupon = headers.index("Coupon") + 1

    values = [sheet.cell(r, coupon).value for r in range(2, sheet.max_row + 1)]
    assert pytest.approx(0.1125) in [v for v in values if v is not None]
    assert sheet.cell(2, coupon).number_format.endswith("%")


def test_undisclosed_non_accrual_is_labelled_not_left_blank(conn, tmp_path):
    """A blank cell reads as 'no'. The workbook has to say 'Not disclosed'."""
    target = tmp_path / "marks.xlsx"
    excel.export_workbook(conn, target)
    sheet = load_workbook(target)["Marks"]
    headers = [cell.value for cell in sheet[1]]
    column = headers.index("Non-accrual") + 1
    values = {sheet.cell(r, column).value for r in range(2, sheet.max_row + 1)}
    assert values == {"Yes", "Not disclosed"}


def test_readme_warns_loudly_when_the_data_is_synthetic(tmp_path):
    connection = db.connect(":memory:")
    db.init_schema(connection)
    db.upsert_bdcs(connection, [ARCC])
    fake = _position(ARCC.cik, "Demo Co, First Lien Term Loan", 1_000, 1_000)
    fake.source = "demo"
    db.load_positions(connection, [fake])

    target = tmp_path / "demo.xlsx"
    result = excel.export_workbook(connection, target)
    assert result["synthetic"] is True

    text = " ".join(
        str(cell.value)
        for row in load_workbook(target)["Read me"].iter_rows()
        for cell in row if cell.value
    )
    assert "THIS FILE CONTAINS SYNTHETIC DATA" in text
    assert "Nothing here was extracted from an SEC filing" in text
    assert "Do not use for analysis" in text
    connection.close()


def test_readme_stays_quiet_when_the_data_is_real(conn, tmp_path):
    target = tmp_path / "real.xlsx"
    excel.export_workbook(conn, target)
    text = " ".join(
        str(cell.value)
        for row in load_workbook(target)["Read me"].iter_rows()
        for cell in row if cell.value
    )
    assert "SYNTHETIC" not in text
    assert "dera" in text


def test_empty_database_is_refused_rather_than_shipping_a_blank_book(tmp_path):
    connection = db.connect(":memory:")
    db.init_schema(connection)
    with pytest.raises(ValueError, match="no marks"):
        excel.export_workbook(connection, tmp_path / "empty.xlsx")
    connection.close()


# ---------------------------------------------------------------------------
# Standalone page
# ---------------------------------------------------------------------------

def test_standalone_page_carries_its_data_and_no_document_scaffolding(conn, tmp_path):
    html = standalone.build_html(conn)
    assert "<title>BDC Tracker</title>" in html
    # An artifact host supplies the scaffolding; a second copy breaks the page.
    for tag in ("<!doctype", "<html", "<head>", "<body"):
        assert tag not in html.lower()
    assert "window.__BDC_BUNDLE__" in html
    assert 'src="app.js"' not in html and 'href="styles.css"' not in html


def test_standalone_bundle_holds_every_view_and_every_bdc(conn):
    bundle = standalone.build_bundle(conn)
    assert {"overview", "bdcs", "disagreements", "meta", "positions"} <= set(bundle)
    assert set(bundle["positions"]) == set(bundle["meta"]["tickers"])
    assert bundle["positions"]["ARCC"]


def test_standalone_escapes_a_closing_script_tag_in_the_data(conn):
    """A borrower name containing </script> would otherwise end the tag early."""
    db.load_positions(conn, [
        _position(ARCC.cik, "</script> Capital, First Lien Term Loan", 1_000, 1_000)
    ])
    conn.commit()
    html = standalone.build_html(conn)
    payload = re.search(r"window\.__BDC_BUNDLE__ = (.*?);</script>", html, re.S).group(1)
    assert "</script>" not in payload
    assert json.loads(payload.replace("<\\/", "</"))


# ---------------------------------------------------------------------------
# Formula correctness — evaluated independently, checked against SQL
# ---------------------------------------------------------------------------
#
# Evaluated with the `formulas` engine rather than LibreOffice: the workbook is
# only trustworthy if its formulas produce the same numbers the database does,
# and that has to be checked by something other than the code that wrote them.


def _evaluate(path) -> dict:
    """Return {SHEET!CELL: value} for every computed cell in the workbook."""
    formulas = pytest.importorskip("formulas")
    import warnings

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        solution = formulas.ExcelModel().loads(str(path)).finish().calculate()

    # Keys look like "'[book.xlsx]BDC SUMMARY'!A2"; reduce to "BDC SUMMARY!A2".
    out = {}
    for key, ranges in solution.items():
        name = key.upper().split("]", 1)[-1].replace("'", "")
        try:
            value = ranges.value[0, 0]
        except Exception:
            continue
        out[name] = value
    return out


def _column(sheet, header: str) -> int:
    return [cell.value for cell in sheet[1]].index(header) + 1


@pytest.mark.slow
def test_row_formulas_compute_the_mark_the_database_computed(conn, tmp_path):
    target = tmp_path / "recalc.xlsx"
    excel.export_workbook(conn, target)
    values = _evaluate(target)

    sheet = load_workbook(target)["Marks"]
    borrower = _column(sheet, "Borrower")
    instrument = _column(sheet, "Instrument")
    mark = get_column_letter(_column(sheet, "Mark (% of principal)"))
    unrealised = get_column_letter(_column(sheet, "Unrealised"))

    rows = {
        (sheet.cell(r, borrower).value, sheet.cell(r, instrument).value): r
        for r in range(2, sheet.max_row + 1)
    }

    acme = rows[("ACME Holdings Inc.", "FIRST_LIEN")]
    assert float(values[f"MARKS!{mark}{acme}"]) == pytest.approx(70.0)

    # Equity has no par, so it must mark against cost: 100k on 500k.
    equity = rows[("Beta Industries", "COMMON_EQUITY")]
    assert float(values[f"MARKS!{mark}{equity}"]) == pytest.approx(20.0)

    beta = rows[("Beta Industries", "SECOND_LIEN")]
    assert float(values[f"MARKS!{unrealised}{beta}"]) == pytest.approx(-3_000_000)


@pytest.mark.slow
def test_summary_totals_match_the_same_query_run_against_sqlite(conn, tmp_path):
    target = tmp_path / "summary.xlsx"
    excel.export_workbook(conn, target)
    values = _evaluate(target)

    sheet = load_workbook(target)["BDC summary"]
    expected = {row["ticker"]: row for row in analytics.bdc_summary(conn, "2025-12-31")}
    rows = {
        sheet.cell(r, 1).value: r
        for r in range(2, sheet.max_row + 1)
        if sheet.cell(r, 1).value in expected
    }
    assert set(rows) == set(expected)

    for ticker, row in rows.items():
        want = expected[ticker]
        assert float(values[f"BDC SUMMARY!G{row}"]) == pytest.approx(want["portfolio_mark"])
        assert float(values[f"BDC SUMMARY!F{row}"]) == pytest.approx(want["fair_value"])
        assert float(values[f"BDC SUMMARY!C{row}"]) == want["positions"]

        # Blank, not 0%, where the filing disclosed no non-accrual status.
        cell = values[f"BDC SUMMARY!K{row}"]
        if want["nonaccrual_pct_fv"] is None:
            assert cell == "", f"{ticker} reported {cell!r} for undisclosed non-accrual"
        else:
            assert float(cell) * 100 == pytest.approx(want["nonaccrual_pct_fv"])
        assert float(values[f"BDC SUMMARY!L{row}"]) == want["nonaccrual_coverage"]


@pytest.mark.slow
def test_summary_ranges_cover_only_their_own_quarter(conn, tmp_path):
    """A prior quarter's rows must not leak into this quarter's totals."""
    earlier = normalize.finalize(
        Position(
            cik=ARCC.cik, period_end=date(2025, 9, 30), source="dera",
            identifier="Acme Holdings, LLC, First Lien Term Loan",
            fair_value=Decimal("9900000"), principal=Decimal("10000000"),
            cost=Decimal("10000000"), accession="0001-25-000000", form="10-Q",
        )
    )
    db.load_positions(conn, [earlier])
    conn.commit()

    target = tmp_path / "twoquarters.xlsx"
    excel.export_workbook(conn, target)
    values = _evaluate(target)

    sheet = load_workbook(target)["BDC summary"]
    rows = {sheet.cell(r, 1).value: r for r in range(2, sheet.max_row + 1)}
    # Still the Q4 figure, not Q4 plus the Q3 row.
    assert float(values[f"BDC SUMMARY!F{rows['ARCC']}"]) == pytest.approx(11_100_000)


@pytest.mark.slow
def test_vintage_sheet_separates_disclosed_years_from_unknown(conn, tmp_path):
    """Untagged positions get their own row, not a made-up year."""
    db.load_positions(conn, [
        normalize.finalize(Position(
            cik=ARCC.cik, period_end=Q2, source="dera",
            identifier="Vintage Co, First Lien Term Loan",
            fair_value=Decimal("900"), principal=Decimal("1000"), cost=Decimal("1000"),
            acquisition_date=date(2021, 5, 1), accession="a", form="10-K",
        ))
    ])
    conn.commit()

    target = tmp_path / "vintage.xlsx"
    excel.export_workbook(conn, target)
    sheet = load_workbook(target)["Vintage & maturity"]
    labels = [sheet.cell(r, 1).value for r in range(1, sheet.max_row + 1)]

    assert 2021 in labels
    assert "Not disclosed" in labels
